from openpyxl import *


def get():
    wb = load_workbook('program音色表.xlsx')

    ws = wb.active
    # 所有行
    keys = []
    values = []
    for row in ws.iter_rows():
        for cell in row:
            # print(cell.value)
            try:
                keys.append(int(cell.value))
            except ValueError:
                values.append(cell.value)
    # # 所有列
    # for column in ws.iter_cols():
    #     for cell in column:
    #         print(cell.value)
    out = ""
    index = 0
    for i in keys:
        out += ", \"" + str(i) + "\": \"" + values[index] + "\""
        index += 1

    print(out)


idList = {"0": "harp", "1": "harp", "2": "pling", "3": "harp", "4": "pling", "5": "pling", "6": "harp", "7": "harp",
          "8": "share", "9": "harp", "10": "didgeridoo", "11": "harp", "12": "xylophone", "13": "chime", "14": "harp",
          "15": "harp", "16": "bass", "17": "harp", "18": "harp", "19": "harp", "20": "harp", "21": "harp",
          "22": "harp",
          "23": "guitar", "24": "guitar", "25": "guitar", "26": "guitar", "27": "guitar", "28": "guitar",
          "29": "guitar",
          "30": "guitar", "31": "bass", "32": "bass", "33": "bass", "34": "bass", "35": "bass", "36": "bass",
          "37": "bass",
          "38": "bass", "39": "bass", "40": "harp", "41": "harp", "42": "harp", "43": "harp", "44": "iron_xylophone",
          "45": "guitar", "46": "harp", "47": "harp", "48": "guitar", "49": "guitar", "50": "bit", "51": "bit",
          "52": "harp", "53": "harp", "54": "bit", "55": "flute", "56": "flute", "57": "flute", "58": "flute",
          "59": "flute", "60": "flute", "61": "flute", "62": "flute", "63": "flute", "64": "bit", "65": "bit",
          "66": "bit", "67": "bit", "68": "flute", "69": "harp", "70": "harp", "71": "flute", "72": "flute",
          "73": "flute", "74": "harp", "75": "flute", "76": "harp", "77": "harp", "78": "harp", "79": "harp",
          "80": "bit", "81": "bit", "82": "bit", "83": "bit", "84": "bit", "85": "bit", "86": "bit", "87": "bit",
          "88": "bit", "89": "bit", "90": "bit", "91": "bit", "92": "bit", "93": "bit", "94": "bit", "95": "bit",
          "96": "bit", "97": "bit", "98": "bit", "99": "bit", "100": "bit", "101": "bit", "102": "bit", "103": "bit",
          "104": "harp", "105": "banjo", "106": "harp", "107": "harp", "108": "harp", "109": "harp", "110": "harp",
          "111": "guitar", "112": "harp", "113": "bell", "114": "harp", "115": "cow_bell", "116": "basedrum",
          "117": "bass", "118": "bit", "119": "basedrum", "120": "guitar", "121": "harp", "122": "harp", "123": "harp",
          "124": "harp", "125": "hat", "126": "basedrum", "127": "snare"}
